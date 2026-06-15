import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from datetime import date, timedelta, datetime
from utils import to_iso, to_display, app_font


class ReportModule(ctk.CTkFrame):
    def __init__(self, parent, db, app=None):
        super().__init__(parent, corner_radius=0, fg_color="transparent")
        self.db = db
        self.app = app
        self._build()

    def _build(self):
        header = ctk.CTkFrame(self, fg_color="white", corner_radius=8)
        header.pack(fill="x", padx=15, pady=(15, 5))
        ctk.CTkLabel(header, text="📊 รายงาน",
                     font=app_font(18, "bold")).pack(side="left", padx=15, pady=10)

        ff = ctk.CTkFrame(self, fg_color="white", corner_radius=8)
        ff.pack(fill="x", padx=15, pady=5)

        row1 = ctk.CTkFrame(ff, fg_color="transparent")
        row1.pack(fill="x")
        row2 = ctk.CTkFrame(ff, fg_color="transparent")
        row2.pack(fill="x")

        ctk.CTkLabel(row1, text="ช่วงวันที่:").pack(side="left", padx=10, pady=8)
        today = date.today()
        first = today.replace(day=1)

        self.date_from = DateEntry(row1, date_pattern="dd/mm/yyyy",
                                    font=("TH Sarabun PSK", 14),
                                    background="#1a5276", foreground="white",
                                    width=10)
        self.date_from.set_date(first)
        self.date_from.pack(side="left", padx=3)

        ctk.CTkLabel(row1, text="—").pack(side="left")

        self.date_to = DateEntry(row1, date_pattern="dd/mm/yyyy",
                                  font=("TH Sarabun PSK", 14),
                                  background="#1a5276", foreground="white",
                                  width=10)
        self.date_to.set_date(today)
        self.date_to.pack(side="left", padx=3)

        ctk.CTkButton(row1, text="🔍 ดูรายงาน", command=self._load,
                      width=110).pack(side="left", padx=8)

        quick_btns = [("วันนี้", 0), ("7 วัน", 7), ("30 วัน", 30), ("เดือนนี้", -1),
                      ("ปีนี้", -2), ("ทั้งหมด", -3)]
        for label, days in quick_btns:
            ctk.CTkButton(row2, text=label, width=72,
                          command=lambda d=days: self._quick_range(d)).pack(side="left", padx=2, pady=8)

        ctk.CTkButton(row2, text="📥 Export Excel", command=self._export,
                      width=130, fg_color="#27ae60", hover_color="#1e8449").pack(side="right", padx=10)
        ctk.CTkButton(row2, text="📄 รายการใช้ยา PDF", command=self._export_pdf_medicine,
                      width=150, fg_color="#c0392b", hover_color="#922b21").pack(side="right", padx=5)
        ctk.CTkButton(row2, text="📄 สถิติ PDF", command=self._export_pdf_stats,
                      width=120, fg_color="#c0392b", hover_color="#922b21").pack(side="right", padx=5)

        # Stat cards
        stat_frame = ctk.CTkFrame(self, fg_color="transparent")
        stat_frame.pack(fill="x", padx=15, pady=5)
        self.stats = {}
        card_defs = [
            ("visits",   "จำนวนครั้งรักษา",   "#2980b9"),
            ("patients", "ผู้ป่วยไม่ซ้ำ",      "#8e44ad"),
            ("avg",      "เฉลี่ย/วัน",         "#27ae60"),
            ("med_types","ชนิดยาที่ใช้",       "#e67e22"),
        ]
        for key, label, color in card_defs:
            card = ctk.CTkFrame(stat_frame, fg_color=color, corner_radius=10)
            card.pack(side="left", fill="both", expand=True, padx=5)
            ctk.CTkLabel(card, text=label, text_color="white",
                         font=app_font(14)).pack(pady=(12, 2))
            val = ctk.CTkLabel(card, text="—", text_color="white",
                               font=app_font(26, "bold"))
            val.pack(pady=(0, 12))
            self.stats[key] = val

        # Two tables
        tables = ctk.CTkFrame(self, fg_color="transparent")
        tables.pack(fill="both", expand=True, padx=15, pady=(5, 15))

        left = ctk.CTkFrame(tables, fg_color="white", corner_radius=8)
        left.pack(side="left", fill="both", expand=True, padx=(0, 5))
        ctk.CTkLabel(left, text="โรค/วินิจฉัยที่พบบ่อย",
                     font=app_font(15, "bold")).pack(anchor="w", padx=15, pady=10)
        self.diag_tree = ttk.Treeview(left, columns=("diag", "cnt"), show="headings", height=14)
        self.diag_tree.heading("diag", text="วินิจฉัย")
        self.diag_tree.heading("cnt",  text="จำนวน")
        self.diag_tree.column("diag", width=230)
        self.diag_tree.column("cnt",  width=70, anchor="center")
        diag_sb = ttk.Scrollbar(left, orient="vertical", command=self.diag_tree.yview)
        self.diag_tree.configure(yscrollcommand=lambda f, l: self._autoscroll(diag_sb, f, l))
        self.diag_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=(0, 10))
        diag_sb.pack(side="right", fill="y", pady=(0, 10))

        right = ctk.CTkFrame(tables, fg_color="white", corner_radius=8)
        right.pack(side="right", fill="both", expand=True, padx=(5, 0))
        ctk.CTkLabel(right, text="ยาที่ใช้บ่อย",
                     font=app_font(15, "bold")).pack(anchor="w", padx=15, pady=10)
        self.med_tree = ttk.Treeview(right, columns=("med", "total"), show="headings", height=14)
        self.med_tree.heading("med",   text="ชื่อยา")
        self.med_tree.heading("total", text="จำนวนรวม")
        self.med_tree.column("med",   width=190)
        self.med_tree.column("total", width=100, anchor="center")
        med_sb = ttk.Scrollbar(right, orient="vertical", command=self.med_tree.yview)
        self.med_tree.configure(yscrollcommand=lambda f, l: self._autoscroll(med_sb, f, l))
        self.med_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=(0, 10))
        med_sb.pack(side="right", fill="y", pady=(0, 10))

        self._load()

    def _autoscroll(self, sb, first, last):
        first, last = float(first), float(last)
        if first <= 0 and last >= 1:
            sb.pack_forget()
        else:
            sb.pack(side="right", fill="y", pady=(0, 10))
        sb.set(first, last)

    def _quick_range(self, days):
        today = date.today()
        if days == -1:
            start = today.replace(day=1)
        elif days == -2:
            start = today.replace(month=1, day=1)
        elif days == -3:
            first = self.db.fetchone("SELECT MIN(visit_date) as d FROM visits")["d"]
            start = datetime.strptime(first, "%Y-%m-%d").date() if first else today
        elif days == 0:
            start = today
        else:
            start = today - timedelta(days=days)
        self.date_from.set_date(start)
        self.date_to.set_date(today)
        self._load()

    def _get_range(self):
        return to_iso(self.date_from.get()), to_iso(self.date_to.get())

    def _load(self):
        d1, d2 = self._get_range()

        total = self.db.fetchone(
            "SELECT COUNT(*) as c FROM visits WHERE visit_date BETWEEN ? AND ?", (d1, d2))["c"]
        unique = self.db.fetchone(
            "SELECT COUNT(DISTINCT emp_id) as c FROM visits WHERE visit_date BETWEEN ? AND ?", (d1, d2))["c"]

        try:
            days = max((datetime.strptime(d2, "%Y-%m-%d") - datetime.strptime(d1, "%Y-%m-%d")).days + 1, 1)
            avg = round(total / days, 1)
        except Exception:
            avg = "—"

        med_types = self.db.fetchone("""
            SELECT COUNT(DISTINCT md.medicine_id) as c
            FROM medicine_dispensed md JOIN visits v ON md.visit_id=v.id
            WHERE v.visit_date BETWEEN ? AND ?
        """, (d1, d2))["c"]

        self.stats["visits"].configure(text=str(total))
        self.stats["patients"].configure(text=str(unique))
        self.stats["avg"].configure(text=str(avg))
        self.stats["med_types"].configure(text=str(med_types))

        diags = self.db.fetchall("""
            SELECT diagnosis, COUNT(*) as c FROM visits
            WHERE visit_date BETWEEN ? AND ? AND diagnosis != ''
            GROUP BY diagnosis ORDER BY c DESC LIMIT 15
        """, (d1, d2))
        self.diag_tree.delete(*self.diag_tree.get_children())
        for r in diags:
            self.diag_tree.insert("", "end", values=(r["diagnosis"], r["c"]))

        meds = self.db.fetchall("""
            SELECT m.medicine_name, SUM(md.quantity) as total
            FROM medicine_dispensed md
            JOIN medicines m ON md.medicine_id=m.id
            JOIN visits v ON md.visit_id=v.id
            WHERE v.visit_date BETWEEN ? AND ?
            GROUP BY m.id ORDER BY total DESC LIMIT 15
        """, (d1, d2))
        self.med_tree.delete(*self.med_tree.get_children())
        for r in meds:
            self.med_tree.insert("", "end", values=(r["medicine_name"], r["total"]))

    def _export(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("ผิดพลาด",
                                 "กรุณาติดตั้ง openpyxl ก่อน:\n\npip install openpyxl")
            return

        fp = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="บันทึกรายงาน",
            initialfile=f"MedicStation_Report_{date.today()}.xlsx")
        if not fp:
            return

        d1, d2 = self._get_range()
        wb = openpyxl.Workbook()
        hf = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="1A5276")
        center = Alignment(horizontal="center")

        def write_sheet(ws, headers, rows):
            for c, h in enumerate(headers, 1):
                cell = ws.cell(1, c, h)
                cell.font = hf
                cell.fill = hfill
                cell.alignment = center
            for r_idx, row in enumerate(rows, 2):
                for c_idx, val in enumerate(row, 1):
                    ws.cell(r_idx, c_idx, val)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 2, 40)

        # Sheet 1 — Summary
        ws0 = wb.active
        ws0.title = "สรุปสถิติ"
        write_sheet(ws0, ["รายการ", "ค่า"], [
            ("ช่วงวันที่", f"{self.date_from.get()} - {self.date_to.get()}"),
            ("จำนวนครั้งรักษา", self.stats["visits"].cget("text")),
            ("ผู้ป่วยไม่ซ้ำ", self.stats["patients"].cget("text")),
            ("เฉลี่ย/วัน", self.stats["avg"].cget("text")),
            ("ชนิดยาที่ใช้", self.stats["med_types"].cget("text")),
        ])

        # Sheet 2 — Visit history
        rows = self.db.fetchall("""
            SELECT v.visit_date, v.emp_id, TRIM(e.rank||' '||e.first_name||' '||e.last_name), e.department,
                   v.chief_complaint, v.diagnosis, v.treatment, v.notes, v.caregiver
            FROM visits v JOIN employees e ON v.emp_id=e.emp_id
            WHERE v.visit_date BETWEEN ? AND ?
            ORDER BY v.visit_date, v.id
        """, (d1, d2))
        write_sheet(wb.create_sheet("ประวัติการรักษา"),
                     ["วันที่", "รหัสพนักงาน", "ชื่อ-นามสกุล", "สังกัด",
                      "อาการ", "วินิจฉัย", "การรักษา", "หมายเหตุ", "ผู้ให้การดูแล"], rows)

        # Sheet 3 — Medicine usage
        med_rows = self.db.fetchall("""
            SELECT m.medicine_name, SUM(md.quantity), COUNT(*)
            FROM medicine_dispensed md
            JOIN medicines m ON md.medicine_id=m.id
            JOIN visits v ON md.visit_id=v.id
            WHERE v.visit_date BETWEEN ? AND ?
            GROUP BY m.id ORDER BY SUM(md.quantity) DESC
        """, (d1, d2))
        write_sheet(wb.create_sheet("การใช้ยา"),
                     ["ชื่อยา", "จำนวนรวม", "จำนวนครั้ง"], med_rows)

        # Sheet 4 — Diagnosis frequency
        diag_rows = self.db.fetchall("""
            SELECT diagnosis, COUNT(*) as c FROM visits
            WHERE visit_date BETWEEN ? AND ? AND diagnosis != ''
            GROUP BY diagnosis ORDER BY c DESC
        """, (d1, d2))
        write_sheet(wb.create_sheet("โรค-วินิจฉัย"), ["วินิจฉัย", "จำนวน"], diag_rows)

        # Sheet 5 — Employees (master list)
        emp_rows = self.db.fetchall("""
            SELECT mn_number, emp_id, TRIM(rank||' '||first_name||' '||last_name), department, gender, blood_type,
                   allergies, underlying_diseases, phone, created_at
            FROM employees ORDER BY emp_id
        """)
        write_sheet(wb.create_sheet("รายชื่อพนักงาน"),
                     ["เลขทะเบียน MN", "รหัสพนักงาน", "ชื่อ-นามสกุล", "สังกัด", "เพศ", "กรุ๊ปเลือด",
                      "ประวัติแพ้ยา", "โรคประจำตัว", "โทรศัพท์", "วันที่ลงทะเบียน"], emp_rows)

        # Sheet 6 — Medicine stock (current)
        stock_rows = self.db.fetchall("""
            SELECT medicine_code, medicine_name, unit, stock_quantity, min_stock, updated_at
            FROM medicines ORDER BY medicine_name
        """)
        write_sheet(wb.create_sheet("คลังยา"),
                     ["รหัสยา", "ชื่อยา", "หน่วย", "คงเหลือ", "ขั้นต่ำ", "อัปเดตล่าสุด"], stock_rows)

        wb.save(fp)
        messagebox.showinfo("สำเร็จ", f"บันทึกรายงานเรียบร้อย\n{fp}")

    def _pdf_prepare(self, file_suffix, title_initial, filename=None):
        """Common setup: import reportlab, register fonts, ask save path.
        Returns (fp, mm, Paragraph, ParagraphStyle, styles) or None if cancelled/failed."""
        try:
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.platypus import Paragraph
            from reportlab.lib.styles import ParagraphStyle
        except ImportError:
            messagebox.showerror("ผิดพลาด",
                                 "กรุณาติดตั้ง reportlab ก่อน:\n\npip install reportlab")
            return None

        import os
        font_dir = r"C:\Windows\Fonts"
        try:
            pdfmetrics.registerFont(TTFont("Tahoma", os.path.join(font_dir, "tahoma.ttf")))
            pdfmetrics.registerFont(TTFont("Tahoma-Bold", os.path.join(font_dir, "tahomabd.ttf")))
        except Exception:
            messagebox.showerror("ผิดพลาด", "ไม่พบฟอนต์ Tahoma สำหรับแสดงผลภาษาไทย")
            return None

        fp = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="บันทึกรายงาน PDF",
            initialfile=filename or f"MedicStation_{file_suffix}_{date.today()}.pdf")
        if not fp:
            return None

        title_style = ParagraphStyle("title", fontName="Tahoma-Bold", fontSize=16, leading=20,
                                       alignment=1, spaceAfter=4)
        sub_style = ParagraphStyle("sub", fontName="Tahoma", fontSize=11, leading=14,
                                     alignment=1, spaceAfter=10)
        heading_style = ParagraphStyle("heading", fontName="Tahoma-Bold", fontSize=12, leading=16,
                                         spaceBefore=12, spaceAfter=6)
        cell_style = ParagraphStyle("cell", fontName="Tahoma", fontSize=9, leading=12)
        footer_style = ParagraphStyle("footer", fontName="Tahoma", fontSize=9, leading=12,
                                        alignment=2, spaceBefore=16)

        elements = [
            Paragraph(title_initial, title_style),
            Paragraph(f"ช่วงวันที่: {self.date_from.get()} - {self.date_to.get()}", sub_style),
        ]
        styles = {"title": title_style, "sub": sub_style, "heading": heading_style,
                  "cell": cell_style, "footer": footer_style}
        return fp, mm, Paragraph, styles, elements

    def _pdf_visit_detail_elements(self, d1, d2, Paragraph, styles, mm):
        """รายละเอียดการรักษา — date, patient, diagnosis, medicines given."""
        cell_style = styles["cell"]
        visit_rows = self.db.fetchall("""
            SELECT v.id, v.visit_date, TRIM(e.rank||' '||e.first_name||' '||e.last_name) as pname,
                   v.diagnosis, v.chief_complaint
            FROM visits v JOIN employees e ON v.emp_id=e.emp_id
            WHERE v.visit_date BETWEEN ? AND ?
            ORDER BY v.visit_date, v.id
        """, (d1, d2))

        med_by_visit = {}
        for r in self.db.fetchall("""
            SELECT md.visit_id, m.medicine_name, md.quantity, m.unit
            FROM medicine_dispensed md
            JOIN medicines m ON md.medicine_id=m.id
            JOIN visits v ON md.visit_id=v.id
            WHERE v.visit_date BETWEEN ? AND ?
            ORDER BY md.id
        """, (d1, d2)):
            qty = r["quantity"]
            qty_str = str(int(qty)) if qty == int(qty) else str(qty)
            med_by_visit.setdefault(r["visit_id"], []).append(
                f"{r['medicine_name']} x{qty_str} {r['unit']}")

        detail_data = [["วันที่", "ผู้ป่วย", "วินิจฉัย/อาการ", "ยาที่ได้รับ"]]
        for r in visit_rows:
            meds = med_by_visit.get(r["id"])
            meds_text = "<br/>".join(meds) if meds else "-"
            diag = r["diagnosis"] or r["chief_complaint"] or "-"
            detail_data.append([
                to_display(r["visit_date"]),
                Paragraph(r["pname"], cell_style),
                Paragraph(diag, cell_style),
                Paragraph(meds_text, cell_style),
            ])
        return [
            Paragraph("รายละเอียดการรักษา", styles["heading"]),
            self._pdf_table(detail_data, [22 * mm, 35 * mm, 43 * mm, 60 * mm]),
        ]

    def _export_pdf_stats(self):
        """สถิติ PDF — สรุปสถิติ + โรค/วินิจฉัยที่พบบ่อย + ยาที่ใช้บ่อย + รายละเอียดการรักษา"""
        prep = self._pdf_prepare("Stats", "รายงานสถิติห้องพยาบาล (Medic Station)")
        if not prep:
            return
        fp, mm, Paragraph, styles, elements = prep
        heading_style = styles["heading"]
        d1, d2 = self._get_range()

        summary_data = [
            ["รายการ", "ค่า"],
            ["จำนวนครั้งรักษา", self.stats["visits"].cget("text")],
            ["ผู้ป่วยไม่ซ้ำ", self.stats["patients"].cget("text")],
            ["เฉลี่ย/วัน", self.stats["avg"].cget("text")],
            ["ชนิดยาที่ใช้", self.stats["med_types"].cget("text")],
        ]
        elements.append(Paragraph("สรุปสถิติ", heading_style))
        elements.append(self._pdf_table(summary_data, [80 * mm, 40 * mm]))

        diag_rows = self.db.fetchall("""
            SELECT diagnosis, COUNT(*) as c FROM visits
            WHERE visit_date BETWEEN ? AND ? AND diagnosis != ''
            GROUP BY diagnosis ORDER BY c DESC LIMIT 20
        """, (d1, d2))
        diag_data = [["วินิจฉัย", "จำนวน"]] + [[r["diagnosis"], str(r["c"])] for r in diag_rows]
        elements.append(Paragraph("โรค/วินิจฉัยที่พบบ่อย", heading_style))
        elements.append(self._pdf_table(diag_data, [100 * mm, 30 * mm]))

        med_rows = self.db.fetchall("""
            SELECT m.medicine_name, SUM(md.quantity) as total
            FROM medicine_dispensed md
            JOIN medicines m ON md.medicine_id=m.id
            JOIN visits v ON md.visit_id=v.id
            WHERE v.visit_date BETWEEN ? AND ?
            GROUP BY m.id ORDER BY total DESC LIMIT 20
        """, (d1, d2))
        med_data = [["ชื่อยา", "จำนวนรวม"]] + [[r["medicine_name"], str(r["total"])] for r in med_rows]
        elements.append(Paragraph("ยาที่ใช้บ่อย", heading_style))
        elements.append(self._pdf_table(med_data, [100 * mm, 30 * mm]))

        elements += self._pdf_visit_detail_elements(d1, d2, Paragraph, styles, mm)

        self._pdf_build(fp, elements, mm)

    def _export_pdf_medicine(self):
        """รายการใช้ยา PDF — สรุปยาที่ใช้ + รายละเอียดการรักษา"""
        prep = self._pdf_prepare("MedicineUsage", "รายงานการใช้ยา แผนกแพทย์ หน่วยบินเดโชชัย 3 ดอนเมือง",
                                  filename=f"รายงานการใช้ยา_เดโชชัย 3_{date.today()}.pdf")
        if not prep:
            return
        fp, mm, Paragraph, styles, elements = prep
        heading_style = styles["heading"]
        d1, d2 = self._get_range()

        med_rows = self.db.fetchall("""
            SELECT m.medicine_name, SUM(md.quantity) as total, COUNT(*) as cnt
            FROM medicine_dispensed md
            JOIN medicines m ON md.medicine_id=m.id
            JOIN visits v ON md.visit_id=v.id
            WHERE v.visit_date BETWEEN ? AND ?
            GROUP BY m.id ORDER BY total DESC
        """, (d1, d2))
        med_data = [["ชื่อยา", "จำนวนรวม", "จำนวนครั้ง"]] + \
            [[r["medicine_name"], str(r["total"]), str(r["cnt"])] for r in med_rows]
        elements.append(Paragraph("สรุปยาที่ใช้", heading_style))
        elements.append(self._pdf_table(med_data, [80 * mm, 35 * mm, 35 * mm]))

        elements += self._pdf_visit_detail_elements(d1, d2, Paragraph, styles, mm)

        exporter = self.app.current_user["display_name"] if self.app and self.app.current_user else "-"
        elements.append(Paragraph(f"ผู้ออกรายงาน: {exporter}", styles["footer"]))

        self._pdf_build(fp, elements, mm)

    def _pdf_build(self, fp, elements, mm):
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate
        doc = SimpleDocTemplate(fp, pagesize=A4,
                                topMargin=20 * mm, bottomMargin=20 * mm,
                                leftMargin=20 * mm, rightMargin=20 * mm)
        doc.build(elements)
        messagebox.showinfo("สำเร็จ", f"บันทึกรายงาน PDF เรียบร้อย\n{fp}")

    def _pdf_table(self, data, col_widths):
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle, Spacer
        from reportlab.lib.units import mm

        if len(data) == 1:
            data.append(["- ไม่มีข้อมูล -"] + [""] * (len(data[0]) - 1))

        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Tahoma"),
            ("FONTNAME", (0, 0), (-1, 0), "Tahoma-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f7")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return Table([[t]], colWidths=[sum(col_widths)],
                      style=TableStyle([("BOTTOMPADDING", (0, 0), (-1, -1), 8)]))

    def refresh(self):
        self._load()
